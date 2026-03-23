"""QAQC Deficiencies Excel file parser service.

Parses RMS QAQC Deficiencies export files and maps them to Procore Observations.
"""
import io
import re
from datetime import datetime, date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from models.rms import RMSDeficiency, RMSDeficiencyParseResult


class QAQCParser:
    """Parser for RMS QAQC Deficiencies Excel files.

    RMS exports have a specific format:
    - Rows 1-10: Header info (title, project name, report date)
    - Row 11: Column headers
    - Row 12+: Data rows
    - Columns are NOT in sequence (1, 3, 6, 8, 11, 14, 15)
    """

    # Column mapping: header name -> Excel column index (1-based)
    COLUMN_MAP = {
        "item_number": 1,
        "description": 3,
        "location": 6,
        "status": 8,
        "date_issued": 11,
        "age_days": 14,
        "staff": 15,
    }

    # Row where headers appear
    HEADER_ROW = 11

    # Row where data starts
    DATA_START_ROW = 12

    def parse(self, file_bytes: bytes) -> RMSDeficiencyParseResult:
        """Parse QAQC Deficiencies Excel file.

        Args:
            file_bytes: Raw bytes of the Excel file

        Returns:
            RMSDeficiencyParseResult with parsed deficiencies
        """
        errors = []
        warnings = []
        deficiencies = []
        project_name = None
        report_date = None
        locations = set()

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active

            # Extract project info from header rows
            project_name = self._extract_project_name(ws)
            report_date = self._extract_report_date(ws)

            # Parse data rows
            row = self.DATA_START_ROW
            while row <= ws.max_row:
                cell_value = ws.cell(row=row, column=1).value
                if cell_value is None:
                    row += 1
                    continue

                item_number = str(cell_value).strip()
                if not item_number.startswith("QA-"):
                    row += 1
                    continue

                try:
                    # Get description (may span multiple rows)
                    description = self._get_cell_text(ws, row, self.COLUMN_MAP["description"])

                    # Get other fields
                    location = self._get_cell_text(ws, row, self.COLUMN_MAP["location"])
                    status = self._get_cell_text(ws, row, self.COLUMN_MAP["status"])
                    date_issued = self._parse_date(
                        ws.cell(row=row, column=self.COLUMN_MAP["date_issued"]).value
                    )
                    age_days = self._parse_int(
                        ws.cell(row=row, column=self.COLUMN_MAP["age_days"]).value
                    )
                    staff = self._get_cell_text(ws, row, self.COLUMN_MAP["staff"])

                    if not status:
                        status = "Open"

                    deficiency = RMSDeficiency(
                        item_number=item_number,
                        description=description or "",
                        location=location,
                        status=status,
                        date_issued=date_issued,
                        age_days=age_days,
                        staff=staff,
                    )
                    deficiencies.append(deficiency)

                    if location:
                        locations.add(location)

                except Exception as e:
                    errors.append(f"Row {row}: Failed to parse deficiency: {str(e)}")

                row += 1

            wb.close()

        except Exception as e:
            errors.append(f"Failed to parse QAQC file: {str(e)}")

        # Calculate stats
        open_count = sum(1 for d in deficiencies if d.is_open)
        closed_count = len(deficiencies) - open_count

        return RMSDeficiencyParseResult(
            deficiencies=deficiencies,
            project_name=project_name,
            report_date=report_date,
            total_count=len(deficiencies),
            open_count=open_count,
            closed_count=closed_count,
            locations=sorted(locations),
            errors=errors,
            warnings=warnings,
        )

    def _extract_project_name(self, ws) -> Optional[str]:
        """Extract project name from header area."""
        # Project name is typically in row 6, column 5
        value = ws.cell(row=6, column=5).value
        if value:
            # Format: "W912QR24C0035  Dobbins AFB, GA -DBB Security Forces Fac"
            return str(value).strip()
        return None

    def _extract_report_date(self, ws) -> Optional[date]:
        """Extract report date from header area."""
        # Report date is typically in row 4, column 13
        value = ws.cell(row=4, column=13).value
        return self._parse_date(value)

    def _get_cell_text(self, ws, row: int, col: int) -> Optional[str]:
        """Get cell text, handling merged cells."""
        value = ws.cell(row=row, column=col).value
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    def _parse_date(self, value) -> Optional[date]:
        """Parse a date value from Excel."""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        try:
            s = str(value).strip()
            # Try common formats
            for fmt in ["%d %b %Y", "%m/%d/%Y", "%Y-%m-%d", "%b %d, %Y"]:
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass

        return None

    def _parse_int(self, value) -> Optional[int]:
        """Parse an integer value."""
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
